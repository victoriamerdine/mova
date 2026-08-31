#!/usr/bin/env python3
"""Importación reproducible de la biblioteca de ejercicios real
(data/ejercicios_consolidado_TOTAL.xlsx) hacia Supabase.

Corrige el hallazgo de la Auditoría 1 (docs/auditoria-01-repositorio-y-arquitectura.md,
sección D.3): el pipeline anterior no era reproducible, era un proceso manual
de una sola vez. Este script se puede correr de nuevo cuantas veces haga
falta (upsert por canonical_name) si el Excel fuente cambia.

Aplica exactamente las reglas NO destructivas ya decididas en
docs/auditoria-02-planificacion-y-biblioteca.md sección J:

  1. Excluye las 2 filas "Otros / No es ejercicio" (no son ejercicios).
  2. Fusiona el 1 duplicado exacto (misma Categoría+Músculo+Ejercicio).
  3. Fusiona los 11 pares de variantes de mayúsculas/espacio, guardando
     AMBAS grafías en exercise_aliases — nunca se pierde el nombre
     original (CLAUDE.md §5).
  4. NO fusiona nada más — ningún otro posible duplicado semántico
     (ej. "Sentadilla libre" / "Back Squat") se toca automáticamente,
     tal como pide la Auditoría 2.
  5. Categoría → pattern_id si es uno de los 8 patrones reales, o
     exercise_stimulus_types si es uno de los 8 tipos de estímulo
     (separación decidida en Auditoría 3 sección C/J.5).
  6. Músculo → muscle_id directo.

Uso:
    python3 scripts/import_exercise_library.py [--dry-run] [--limit N]

Requiere que scripts/seed_reference_data.py ya se haya corrido antes
(necesita los patterns/muscles/stimulus_types ya cargados para poder
resolver sus id).
"""

import argparse
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _supabase import SupabaseRest  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = REPO_ROOT / "data" / "ejercicios_consolidado_TOTAL.xlsx"

JUNK_CATEGORY = "Otros / No es ejercicio"

# Auditoría 2 sección F: mapeo de "Estado de Coincidencia" (texto de la
# planilla) a los valores del CHECK constraint de exercises.match_status.
MATCH_STATUS_MAP = {
    "Coincidencia exacta": "coincidencia_exacta",
    "Coincidencia probable": "coincidencia_probable",
    "Aproximado (revisar)": "aproximado_revisar",
    "Ambiguo (varias opciones posibles)": "ambiguo",
    "Sin video encontrado": "sin_video_encontrado",
}

SOURCE_MAP = {
    "Base Original": "base_original",
    "Nuevo (Profe)": "nuevo_profe",
}


def normalize_key(name: str) -> str:
    """Clave case/espacio-insensitive para detectar duplicados — misma
    lógica que se usó para encontrarlos en la Auditoría 2."""
    normalized = unicodedata.normalize("NFKC", str(name)).strip().lower()
    return re.sub(r"\s+", " ", normalized)


def read_rows():
    import openpyxl

    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb["Consolidado"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Ejercicio"]] is None:
            continue
        rows.append(
            {
                "categoria": (row[idx["Categoría"]] or "").strip(),
                "musculo": (row[idx["Músculo"]] or "").strip(),
                "ejercicio": str(row[idx["Ejercicio"]]).strip(),
                "links": [
                    row[idx[f"Link {n}"]].strip()
                    for n in range(1, 6)
                    if row[idx[f"Link {n}"]] and str(row[idx[f"Link {n}"]]).strip()
                ],
                "origen": (row[idx["Origen"]] or "").strip(),
                "estado": (row[idx["Estado de Coincidencia"]] or "").strip() or None,
            }
        )
    return rows


def dedupe(rows: list[dict]) -> list[dict]:
    """Agrupa por nombre normalizado. El primero que aparece en el Excel
    queda como canónico; el resto se guardan como aliases (grafía
    original). Reporta lo que hizo, no lo esconde."""
    groups: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        groups[normalize_key(r["ejercicio"])].append(r)

    merged = []
    for _key, group in groups.items():
        canonical = group[0]
        aliases = [g["ejercicio"] for g in group[1:] if g["ejercicio"] != canonical["ejercicio"]]
        if len(group) > 1:
            print(f"  fusionado: {[g['ejercicio'] for g in group]} -> \"{canonical['ejercicio']}\"")
        canonical = dict(canonical)
        canonical["aliases"] = aliases
        merged.append(canonical)
    return merged


def build_exercise_rows(rows, pattern_by_name, muscle_by_name, stimulus_by_name):
    exercises_payload = []
    for r in rows:
        muscle_id = muscle_by_name.get(normalize_key(r["musculo"]))
        pattern_id = pattern_by_name.get(normalize_key(r["categoria"]))
        stimulus_id = stimulus_by_name.get(normalize_key(r["categoria"]))

        match_status = MATCH_STATUS_MAP.get(r["estado"]) if r["estado"] else None
        source = SOURCE_MAP.get(r["origen"], "nuevo_profe")

        exercises_payload.append(
            {
                "canonical_name": r["ejercicio"],
                "display_name": r["ejercicio"],
                "original_name": r["ejercicio"],
                "muscle_id": muscle_id,
                "pattern_id": pattern_id,
                "source": source,
                "match_status": match_status,
                "status": "active",
                "_stimulus_id": stimulus_id,  # se usa después, no es columna real
                "_aliases": r["aliases"],
                "_links": r["links"],
            }
        )
    return exercises_payload


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--limit", type=int, default=None, help="Solo procesar las primeras N filas (para probar rápido)")
    args = parser.parse_args()

    db = SupabaseRest(dry_run=args.dry_run)

    print("=== Leyendo Excel fuente ===")
    rows = read_rows()
    print(f"  {len(rows)} filas totales")

    before = len(rows)
    rows = [r for r in rows if r["categoria"] != JUNK_CATEGORY and r["musculo"] != JUNK_CATEGORY]
    print(f"  excluidas {before - len(rows)} filas 'Otros / No es ejercicio'")

    print("=== Detectando y fusionando duplicados (case/espacio) ===")
    rows = dedupe(rows)
    print(f"  {len(rows)} ejercicios únicos después de fusionar")

    if args.limit:
        rows = rows[: args.limit]
        print(f"  (--limit activo: procesando solo {len(rows)})")

    print("=== Resolviendo catálogos ya sembrados (patterns/muscles/stimulus_types) ===")
    patterns = db.select("patterns", "select=id,canonical_name")
    muscles = db.select("muscles", "select=id,canonical_name")
    stimulus_types = db.select("stimulus_types", "select=id,canonical_name")

    if not args.dry_run and (not patterns or not muscles or not stimulus_types):
        raise SystemExit(
            "Los catálogos de referencia están vacíos. Corré primero "
            "scripts/seed_reference_data.py antes de importar la biblioteca."
        )

    pattern_by_name = {normalize_key(p["canonical_name"]): p["id"] for p in patterns}
    muscle_by_name = {normalize_key(m["canonical_name"]): m["id"] for m in muscles}
    stimulus_by_name = {normalize_key(s["canonical_name"]): s["id"] for s in stimulus_types}

    exercises_payload = build_exercise_rows(rows, pattern_by_name, muscle_by_name, stimulus_by_name)

    unmatched_categoria = {
        r["categoria"]
        for r, e in zip(rows, exercises_payload)
        if r["categoria"] and e["pattern_id"] is None and e["_stimulus_id"] is None
    }
    unmatched_musculo = {r["musculo"] for r in rows if r["musculo"] and normalize_key(r["musculo"]) not in muscle_by_name}
    if unmatched_categoria:
        print(f"  AVISO: categorías sin match en patterns/stimulus_types: {sorted(unmatched_categoria)}")
    if unmatched_musculo:
        print(f"  AVISO: músculos sin match en catálogo: {sorted(unmatched_musculo)}")

    print(f"=== Insertando {len(exercises_payload)} ejercicios ===")
    clean_rows = [
        {k: v for k, v in e.items() if not k.startswith("_")} for e in exercises_payload
    ]
    inserted = db.upsert("exercises", clean_rows, on_conflict="canonical_name")
    id_by_name = {normalize_key(e["canonical_name"]): e["id"] for e in inserted}
    print(f"  OK: {len(inserted)} ejercicios")

    print("=== Insertando aliases ===")
    alias_rows = []
    for e in exercises_payload:
        ex_id = id_by_name.get(normalize_key(e["canonical_name"]))
        if not ex_id:
            continue
        for alias in e["_aliases"]:
            alias_rows.append({"exercise_id": ex_id, "alias": alias, "note": "variante de mayúsculas/espacio fusionada en la importación"})
    if alias_rows:
        db.upsert("exercise_aliases", alias_rows, on_conflict="exercise_id,alias")
    print(f"  OK: {len(alias_rows)} aliases")

    print("=== Insertando relación con tipos de estímulo ===")
    stimulus_rows = []
    for e in exercises_payload:
        if not e["_stimulus_id"]:
            continue
        ex_id = id_by_name.get(normalize_key(e["canonical_name"]))
        if ex_id:
            stimulus_rows.append({"exercise_id": ex_id, "stimulus_type_id": e["_stimulus_id"]})
    if stimulus_rows:
        db.upsert("exercise_stimulus_types", stimulus_rows, on_conflict="exercise_id,stimulus_type_id")
    print(f"  OK: {len(stimulus_rows)} relaciones")

    print("=== Insertando videos (exercise_media) ===")
    media_rows = []
    for e in exercises_payload:
        ex_id = id_by_name.get(normalize_key(e["canonical_name"]))
        if not ex_id:
            continue
        for i, link in enumerate(e["_links"]):
            media_rows.append(
                {
                    "exercise_id": ex_id,
                    "type": "video",
                    "url": link,
                    "source": "youtube",
                    "is_primary": i == 0,
                    "sort_order": i,
                    "status": "active",
                }
            )
    if media_rows:
        db.upsert("exercise_media", media_rows, on_conflict="exercise_id,url")
    print(f"  OK: {len(media_rows)} videos")

    print()
    print("=== Reporte final ===")
    print(f"  ejercicios importados:        {len(inserted)}")
    print(f"  aliases (variantes fusionadas): {len(alias_rows)}")
    print(f"  videos:                        {len(media_rows)}")
    con_video = sum(1 for e in exercises_payload if e['_links'])
    print(f"  ejercicios con al menos 1 video: {con_video}")
    print(f"  ejercicios sin video:            {len(exercises_payload) - con_video}")


if __name__ == "__main__":
    main()
